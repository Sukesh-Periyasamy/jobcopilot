"""Unit tests for the export service module.

Tests cover:
- Empty export (header-only) for CSV and XLSX
- Truncation at 10000 records
- CSV output format correctness
- XLSX output format correctness
- No truncation under the limit

Requirements: 5.4, 5.5, 5.6, 5.7
"""

import csv
import io

import openpyxl
import pytest

from app.models.job import JobRecord
from app.services.export_service import (
    EXPORT_COLUMNS,
    MAX_EXPORT_RECORDS,
    export_csv,
    export_xlsx,
)


def _make_job(index: int) -> JobRecord:
    """Create a sample JobRecord with deterministic values based on index."""
    return JobRecord(
        title=f"Engineer {index}",
        company=f"Company {index}",
        location=f"City {index}",
        source="linkedin",
        job_url=f"https://example.com/job/{index}",
        description=f"Description {index}",
        job_type="Full-time",
        salary=f"${index * 1000}",
        date_posted=f"2024-01-{(index % 28) + 1:02d}",
        search_term="engineer",
        source_type="jobspy",
        source_platform="linkedin",
        created_at="2024-01-15T08:00:00+00:00",
    )


class TestExportCSVEmpty:
    """Test empty export CSV returns header-only file."""

    def test_empty_csv_returns_header_only(self):
        buffer, was_truncated = export_csv([])

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 1  # Only header row
        assert rows[0] == EXPORT_COLUMNS
        assert was_truncated is False

    def test_empty_csv_buffer_at_position_zero(self):
        buffer, _ = export_csv([])
        # Buffer should be seeked to 0 (readable immediately)
        assert buffer.tell() == 0


class TestExportXLSXEmpty:
    """Test empty export XLSX returns header-only file."""

    def test_empty_xlsx_returns_header_only(self):
        buffer, was_truncated = export_xlsx([])

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active

        # Should have only 1 row (header)
        rows = list(ws.iter_rows(values_only=True))
        assert len(rows) == 1
        assert list(rows[0]) == EXPORT_COLUMNS
        assert was_truncated is False

    def test_empty_xlsx_buffer_at_position_zero(self):
        buffer, _ = export_xlsx([])
        assert buffer.tell() == 0


class TestExportCSVTruncation:
    """Test truncation at MAX_EXPORT_RECORDS for CSV."""

    def test_truncation_at_limit(self):
        jobs = [_make_job(i) for i in range(MAX_EXPORT_RECORDS + 1)]

        buffer, was_truncated = export_csv(jobs)

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        # Header + MAX_EXPORT_RECORDS data rows
        assert len(rows) == MAX_EXPORT_RECORDS + 1
        assert was_truncated is True

    def test_no_truncation_under_limit(self):
        jobs = [_make_job(i) for i in range(100)]

        buffer, was_truncated = export_csv(jobs)

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 101  # Header + 100 data rows
        assert was_truncated is False


class TestExportXLSXTruncation:
    """Test truncation at MAX_EXPORT_RECORDS for XLSX."""

    def test_truncation_at_limit(self):
        jobs = [_make_job(i) for i in range(MAX_EXPORT_RECORDS + 1)]

        buffer, was_truncated = export_xlsx(jobs)

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        # Header + MAX_EXPORT_RECORDS data rows
        assert len(rows) == MAX_EXPORT_RECORDS + 1
        assert was_truncated is True

    def test_no_truncation_under_limit(self):
        jobs = [_make_job(i) for i in range(100)]

        buffer, was_truncated = export_xlsx(jobs)

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 101  # Header + 100 data rows
        assert was_truncated is False


class TestExportCSVFormat:
    """Test CSV output has correct columns and data values."""

    def test_csv_columns_match_export_columns(self):
        jobs = [_make_job(0)]

        buffer, _ = export_csv(jobs)

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert rows[0] == EXPORT_COLUMNS

    def test_csv_data_values_correct(self):
        job = _make_job(42)
        buffer, _ = export_csv([job])

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        data_row = rows[1]
        # Verify values match the job record in EXPORT_COLUMNS order
        assert data_row[0] == "Engineer 42"       # title
        assert data_row[1] == "Company 42"        # company
        assert data_row[2] == "City 42"           # location
        assert data_row[3] == "linkedin"          # source
        assert data_row[4] == "jobspy"            # source_type
        assert data_row[5] == "linkedin"          # source_platform
        assert data_row[6] == "https://example.com/job/42"  # job_url
        assert data_row[7] == "Description 42"    # description
        assert data_row[8] == "Full-time"         # job_type
        assert data_row[9] == "$42000"            # salary
        assert data_row[10] == "2024-01-15"       # date_posted
        assert data_row[11] == "engineer"         # search_term
        assert data_row[12] == "2024-01-15T08:00:00+00:00"  # created_at

    def test_csv_multiple_rows(self):
        jobs = [_make_job(i) for i in range(5)]
        buffer, _ = export_csv(jobs)

        content = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)

        assert len(rows) == 6  # Header + 5 data rows
        # Verify each row has the correct number of columns
        for row in rows:
            assert len(row) == len(EXPORT_COLUMNS)


class TestExportXLSXFormat:
    """Test XLSX output has correct columns and data values."""

    def test_xlsx_columns_match_export_columns(self):
        jobs = [_make_job(0)]

        buffer, _ = export_xlsx(jobs)

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert list(rows[0]) == EXPORT_COLUMNS

    def test_xlsx_data_values_correct(self):
        job = _make_job(42)
        buffer, _ = export_xlsx([job])

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        data_row = rows[1]
        assert data_row[0] == "Engineer 42"       # title
        assert data_row[1] == "Company 42"        # company
        assert data_row[2] == "City 42"           # location
        assert data_row[3] == "linkedin"          # source
        assert data_row[4] == "jobspy"            # source_type
        assert data_row[5] == "linkedin"          # source_platform
        assert data_row[6] == "https://example.com/job/42"  # job_url
        assert data_row[7] == "Description 42"    # description
        assert data_row[8] == "Full-time"         # job_type
        assert data_row[9] == "$42000"            # salary
        assert data_row[10] == "2024-01-15"       # date_posted
        assert data_row[11] == "engineer"         # search_term
        assert data_row[12] == "2024-01-15T08:00:00+00:00"  # created_at

    def test_xlsx_multiple_rows(self):
        jobs = [_make_job(i) for i in range(5)]
        buffer, _ = export_xlsx(jobs)

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        assert len(rows) == 6  # Header + 5 data rows
        # Verify each row has the correct number of columns
        for row in rows:
            assert len(row) == len(EXPORT_COLUMNS)

    def test_xlsx_sheet_name(self):
        buffer, _ = export_xlsx([_make_job(0)])

        wb = openpyxl.load_workbook(buffer)
        ws = wb.active
        assert ws.title == "Jobs"


# ---------------------------------------------------------------------------
# Feature: jobcopilot-v1.1-upgrade, Property 7: Export Round-Trip Correctness
# ---------------------------------------------------------------------------

from hypothesis import given, settings as hyp_settings
from hypothesis import strategies as st


# ---------------------------------------------------------------------------
# Hypothesis strategies
# ---------------------------------------------------------------------------

# Simple alphanumeric strings to avoid CSV parsing issues with special characters
_safe_text = st.text(
    alphabet="abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789 ",
    min_size=1,
    max_size=30,
)

_source_type_st = st.sampled_from(["jobspy", "jobhive"])
_source_platform_st = st.sampled_from(["linkedin", "indeed", "greenhouse", "workday"])

# Optional fields: either empty string or a safe alphanumeric string
_optional_text = st.one_of(st.just(""), _safe_text)

# ISO format datetime string strategy
_created_at_st = st.from_regex(
    r"2024-0[1-9]-[012][0-9]T[01][0-9]:[0-5][0-9]:[0-5][0-9]\+00:00",
    fullmatch=True,
)

# Strategy for a single JobRecord
_job_record_st = st.builds(
    JobRecord,
    title=_safe_text,
    company=_safe_text,
    location=_safe_text,
    source=_safe_text,
    job_url=st.from_regex(r"https://example\.com/jobs/[a-z0-9]{5}", fullmatch=True),
    description=_optional_text,
    job_type=_optional_text,
    salary=_optional_text,
    date_posted=_optional_text,
    search_term=_optional_text,
    source_type=_source_type_st,
    source_platform=_source_platform_st,
    created_at=_created_at_st,
)

# Strategy for a list of 1-10 JobRecords
_job_list_st = st.lists(_job_record_st, min_size=1, max_size=10)


# ---------------------------------------------------------------------------
# Property tests
# ---------------------------------------------------------------------------


class TestExportRoundTripCorrectnessProperty:
    """Property test: Export Round-Trip Correctness.

    For any list of JobRecords, exporting to CSV or XLSX and then parsing the
    output SHALL produce rows containing exactly the defined export columns with
    values matching the original records.

    **Validates: Requirements 5.1, 5.2, 5.4**
    """

    @hyp_settings(max_examples=100)
    @given(jobs=_job_list_st)
    def test_csv_round_trip_header_matches_export_columns(self, jobs: list[JobRecord]):
        """CSV export header row matches EXPORT_COLUMNS exactly."""
        buffer, _ = export_csv(jobs)
        text = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)
        assert header == EXPORT_COLUMNS

    @hyp_settings(max_examples=100)
    @given(jobs=_job_list_st)
    def test_csv_round_trip_data_matches_originals(self, jobs: list[JobRecord]):
        """CSV export data rows match the corresponding JobRecord field values."""
        buffer, _ = export_csv(jobs)
        text = buffer.read().decode("utf-8")
        reader = csv.reader(io.StringIO(text))
        header = next(reader)  # skip header
        rows = list(reader)

        assert len(rows) == len(jobs)

        for row, job in zip(rows, jobs):
            for col_idx, col_name in enumerate(EXPORT_COLUMNS):
                expected = getattr(job, col_name, "")
                actual = row[col_idx]
                assert actual == expected, (
                    f"CSV mismatch for column '{col_name}': "
                    f"expected {expected!r}, got {actual!r}"
                )

    @hyp_settings(max_examples=100)
    @given(jobs=_job_list_st)
    def test_xlsx_round_trip_header_matches_export_columns(self, jobs: list[JobRecord]):
        """XLSX export header row matches EXPORT_COLUMNS exactly."""
        buffer, _ = export_xlsx(jobs)
        from openpyxl import load_workbook

        wb = load_workbook(buffer)
        ws = wb.active
        header = [cell.value for cell in ws[1]]
        assert header == EXPORT_COLUMNS

    @hyp_settings(max_examples=100)
    @given(jobs=_job_list_st)
    def test_xlsx_round_trip_data_matches_originals(self, jobs: list[JobRecord]):
        """XLSX export data rows match the corresponding JobRecord field values."""
        buffer, _ = export_xlsx(jobs)
        from openpyxl import load_workbook

        wb = load_workbook(buffer)
        ws = wb.active

        # Data starts at row 2 (row 1 is header)
        data_rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(data_rows) == len(jobs)

        for row, job in zip(data_rows, jobs):
            for col_idx, col_name in enumerate(EXPORT_COLUMNS):
                expected = getattr(job, col_name, "")
                actual = row[col_idx]
                # openpyxl may return None for empty cells
                if actual is None:
                    actual = ""
                assert str(actual) == expected, (
                    f"XLSX mismatch for column '{col_name}': "
                    f"expected {expected!r}, got {actual!r}"
                )
